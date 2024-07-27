from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from linkedin import LinkedIn
import os

class MakeExcelFile():
    """
    A class to create an Excel file from job application data.
    """
    def make_excel(self,applied_jobs,companies_name):
        """
        Create an Excel file with job application data.

        The method performs the following:
        - Merges header cells for better readability.
        - Applies alignment for better readability.
        - Iterates over job application data, merges cells for each job application, and populates them.
        - Saves the workbook as 'job_applications.xlsx'.
        """
        wb = Workbook()
        ws = wb.active

        # Merge header cells
        ws.merge_cells("A1:C1")
        ws.merge_cells("D1:E1")
        ws.merge_cells("F1:I1")
        ws.merge_cells("J1:M1")
        ws.merge_cells("N1:S1")

        # Set the header names
        ws["A1"] = "Company"
        ws["D1"] = "Applied Time"
        ws["F1"] = "Role"
        ws["J1"] = "Location"
        ws["N1"] = "Link"

        # Apply alignment to header cells for center alignment
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
        ws["N1"].alignment = Alignment(horizontal='center', vertical='center')
        
        ind = 2
        try:
            for comp in applied_jobs:
                ws.merge_cells(f"A{ind}:C{ind+len(applied_jobs[comp])-1}")
                ws[f"A{ind}"] = companies_name[comp]
                for detail in applied_jobs[comp]:
                    ws.merge_cells(f"D{ind}:E{ind}")    # Merge 2 columns
                    ws.merge_cells(f"F{ind}:I{ind}")    # Merge 4 columns
                    ws.merge_cells(f"J{ind}:M{ind}")    # Merge 4 columns
                    ws.merge_cells(f"N{ind}:S{ind}")
                    ws[f"D{ind}"] = detail[1]   # applied_date
                    ws[f"F{ind}"] = detail[0]   # role element
                    ws[f"J{ind}"] = detail[2]   # location
                    ws[f"N{ind}"] = detail[3]   # Application link

                    # Apply alignment to the data cells for center alignment
                    ws[f"A{ind}"].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f"D{ind}"].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f"F{ind}"].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f"J{ind}"].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f"N{ind}"].alignment = Alignment(horizontal='center', vertical='center')
                    ind += 1

            wb.save("job_applications.xlsx") # Save the workbook to a file
        except Exception as e:
            print("Exception occured: ", e)
        else:
            print("Excel file has been created.\n")

    def check_excel_existence(self):
        current_directory = os.getcwd() # Get the current working directory
        if "job_applications.xlsx" in os.listdir(current_directory):
            return True
        else:
            return False
    
    def load_excel(self):
        wb = load_workbook('job_applications.xlsx')
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        return data
